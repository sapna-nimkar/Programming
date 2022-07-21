'''
Steps to perform:
1. Make a GET request call to https://reqres.in/api/users?page=2 : to get Users list as json response
2. Create an excel with header columns having same details about the user
3. from Json response, take values and write in the excel in relevant columns
'''

import json
import openpyxl
import requests

url = 'https://reqres.in/api/users?page=1'
headers = {'Connection': 'keep-alive'}

def get_users(url):
    response = requests.get(url, headers = headers)
    header_details = list(response.json()['data'][0].keys())
    print(header_details)

    #create excel and put this header details in the columns header

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'user_details'
    wb.save("./user.xlsx")

    #put header details in the excel
    for col in range(1,len(header_details)+1):
        cell = sheet.cell(row = 1, column = col)
        cell.value = header_details[col-1]

    wb.save("./user.xlsx")

#write user details in the excel rows
    data_values = response.json()['data']
    # for row in range(1, len(data_values)):
    #     for col in range(1,len(header_details)):
    #         cell = sheet.cell(row = row+1, column = col)
    #         for key in data_values[row-1]:

    #             cell.value = data_values[row-1][key]
    data = []
    for item in data_values:
        row = []
        for key in header_details:
            row.append(item[key])
        data.append(row)

    for row in range(1, len(data)+1):
        for col in range(1, len(data[row-1])+1):
            cell = sheet.cell(row=row+1, column = col)
            cell.value = data[row-1][col-1]
        

    wb.save("./user.xlsx")





if __name__ == '__main__':
    url = 'https://reqres.in/api/users?page=2'
    get_users(url)


