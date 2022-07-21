#read and write to an excel file

import openpyxl

#to create a new blank workbook object
wb = openpyxl.Workbook()

#get Active sheet from the workbook
sheet = wb.active

#sheet_title = sheet.title
#print('active sheet title is {}'.format(sheet_title))

#change the title of this sheet
sheet.title = 'sheetOne'
sheet_title = sheet.title
print('active sheet title is {}'.format(sheet_title))

#create a cell object which has rows and columns which start from 1 and not 0
c1 = sheet.cell(row=1, column=1)

#writing values to cell
c1.value = "Sapna"

c2 = sheet.cell(row=1, column=2)
c2.value = "Supratim"

#can also access using actual cell name

c3 = sheet['A2']
c3.value = 'Samantray'


for i in range(1,7):
    values1 = ['Sapna', 'Supratim', 'Joey', 'Zolando', 'Berlin', 'Germany']
    cell = sheet['A{}'.format(i)]
    cell.value = values1[i-1]

    values2 = ['Nimkar', 'Samantray', 'Doggola', 'Company', 'City', 'Country']
    cell = sheet['B{}'.format(i)]
    cell.value = values2[i-1]



#to Add more sheets into the workbook
wb.create_sheet(index=1, title= 'sheetTwo')
wb.active = wb['sheetTwo']
sheet = wb.active
cell = sheet['A1']
cell.value = 'Newsheetcell'

data = [['Sapna', 'Nimkar', '9993882313'], ['Joey', 'Doggola', '9903834200']]

for row in range(1,len(data)+1):
    for col in range(1,len(data[row-1])+1):
        cell = sheet.cell(row=row, column = col)
        cell.value = data[row-1][col-1]


# command to create and save the excel
wb.save("./testexcel.xlsx")

#To READ from Excel
# print value from a particular cell
path = "./testexcel.xlsx"
wb_object = openpyxl.load_workbook(path)
wb_object.active = wb_object["sheetTwo"]
sheet_obj = wb_object.active
cell_obj = sheet_obj.cell(row=2, column = 2)

print(cell_obj.value)




